"""
============================================================================
EXTRAÇÃO DE DADOS EXPERIMENTAIS — TIER 1
Modelo Constitutivo MLP para Areias Reforçadas com Fibras Poliméricas

Autor: João Paulo de Araújo Rodrigues — Mestrado / UFU
Orientador: Prof. Dr. Michael Andrade Maedo

Objetivo:
    Ler as planilhas 'FRS_2002_mod_conceicao_pf05.xlsx' e
    'FRS_2002_mod_conceicao_Pf1.xlsx', extrair os 24 ensaios triaxiais
    drenados (CDC) da aba 'Exp_res' de cada uma, e gerar um único CSV
    consolidado no formato compatível com o pipeline da rede neural MLP.

Saída esperada:
    24 ensaios = 12 (Pf=1%, da Pf1) + 12 (4 sem fibra + 8 com Pf=0,5%, da pf05)
    ~1.200 pontos experimentais

Premissa documentada:
    Todos os corpos de prova foram preparados a Dr = 60% (PINTO, 2021),
    resultando em e0 ≈ 0,613 constante para todos os ensaios. Esse valor
    é o mesmo já adotado pelo João Paulo no artigo do COBRAMSEG.

Formato do CSV de saída (13 colunas, idêntico ao dados_experimentais_processados.csv):
    po_kpa, teor_de_fibra, comp_de_fibra, e_inicial_vazio_inicial,
    p_tensao_media, q_tensao_desviadora, e_indice_de_vazios,
    ea_deformacao_axial, ev_deformacao_volumetrica,
    deltaEa, deltaev, dq, dp
============================================================================
"""

import re
import pandas as pd
from openpyxl import load_workbook

# ============================================================================
# CONFIGURAÇÃO
# ============================================================================

PLANILHAS = {
    "pf05": "FRS_2002_mod_conceicao_pf05.xlsx",
    "Pf1":  "FRS_2002_mod_conceicao_Pf1.xlsx",
}

CSV_SAIDA = "dados_experimentais_tier1.csv"

# Índice de vazios inicial — constante (Dr = 60%, PINTO 2021)
# Mesmo valor usado pelo JP no artigo (dados_experimentais_processados.csv)
E0_CONSTANTE = 0.613

# Limite máximo de colunas para procurar ensaios na Exp_res
# A partir da coluna ~49 começam as simulações numéricas (não interessam)
MAX_COL_EXP = 48

# ============================================================================
# PARSERS DE CABEÇALHO
# ============================================================================

def parse_sigma3(v):
    """Extrai σ3 de strings como '50 KPA', '100 KPA' etc."""
    if v is None:
        return None
    s = str(v).strip().upper()
    m = re.match(r'^(\d+)\s*KPA', s)
    return int(m.group(1)) if m else None


def parse_pf(v):
    """
    Extrai teor de fibra Pf (em %) — valores possíveis:
    - None ou string vazia → 0.0 (sem fibra)
    - 0.005 → 0.5 (formato decimal/100)
    - 0.5 → 0.5 (formato direto)
    - 0.01 → 1.0 (formato decimal/100)
    - 1 → 1.0 (formato direto)
    """
    if v is None or (isinstance(v, str) and not v.strip()):
        return 0.0
    if isinstance(v, str):
        try:
            v = float(v.replace(",", "."))
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        f = float(v)
        if f == 0:
            return 0.0
        if f == 0.005:
            return 0.5
        if f == 0.5:
            return 0.5
        if f == 0.01:
            return 1.0
        if f == 1:
            return 1.0
    return None


def parse_L(v):
    """Extrai comprimento L de 'L25', 'L51', 'L12,5', etc."""
    if v is None:
        return None
    s = str(v).strip().upper()
    m = re.match(r'^L\s*([\d,\.]+)', s)
    return float(m.group(1).replace(",", ".")) if m else None


# ============================================================================
# EXTRAÇÃO DE UM ENSAIO
# ============================================================================

def _detectar_inicio_dados(all_rows, col_inicio):
    """
    Detecta automaticamente em qual linha começam os dados experimentais.

    Layout A: L1=σ3,Pf,L | L2=labels | L3+=dados (primeiro ponto = 0,0,0)
    Layout B: L1=σ3,Pf,L | L2+=dados (primeiro ponto = 0,0,0)

    Retorna o índice (0-based) da primeira linha de dados.
    """
    # A linha 1 (índice 0) sempre é o cabeçalho σ3/Pf/L → não é dados
    # Verificar se a linha 2 (índice 1) é label (string) ou dado (numérico)
    if 1 < len(all_rows):
        v = all_rows[1][col_inicio] if col_inicio < len(all_rows[1]) else None
        if isinstance(v, str):
            return 2  # Layout A: dados começam em L3 (índice 2)
        if isinstance(v, (int, float)):
            return 1  # Layout B: dados começam em L2 (índice 1)
    return 2  # default: Layout A


def extrair_ensaio(all_rows, col_inicio, sigma3, pf, L, max_col_exp):
    """
    Extrai os pontos (εa, q, εv) de um único ensaio,
    a partir da coluna 'col_inicio' (0-indexed).
    Retorna DataFrame ou None se o ensaio não for válido.
    """
    if col_inicio + 2 >= max_col_exp:
        return None

    # Detectar onde começam os dados (Layout A vs Layout B)
    linha_inicio = _detectar_inicio_dados(all_rows, col_inicio)

    # Coletar pontos a partir da linha detectada
    pontos = []
    for row in all_rows[linha_inicio:]:
        if col_inicio + 2 >= len(row):
            break
        ea = row[col_inicio]
        q = row[col_inicio + 1]
        ev = row[col_inicio + 2]
        # Converter strings numéricas (raro mas ocorre)
        for nome, val in [("ea", ea), ("q", q), ("ev", ev)]:
            if isinstance(val, str):
                try:
                    val_f = float(val.replace(",", "."))
                except ValueError:
                    val_f = None
                if nome == "ea": ea = val_f
                elif nome == "q": q = val_f
                elif nome == "ev": ev = val_f
        # Pular ponto se algum dos 3 não for numérico (linhas vazias intercaladas)
        if not all(isinstance(x, (int, float)) for x in [ea, q, ev]):
            continue
        pontos.append({"ea": ea, "q": q, "ev": ev})

    if len(pontos) < 5:
        return None

    df = pd.DataFrame(pontos)

    # Calcular p (tensão média efetiva, ensaio drenado)
    # Em CDC drenado triaxial: σ1 = σ3 + q, p = (σ1 + 2σ3)/3 = σ3 + q/3
    df["p"] = sigma3 + df["q"] / 3.0

    # Calcular índice de vazios atual
    # e_atual = e0 - εv*(1 + e0)   [relação clássica em mecânica dos solos]
    df["e_atual"] = E0_CONSTANTE - df["ev"] * (1 + E0_CONSTANTE)

    # Garantir que o primeiro ponto começa exatamente em (0, 0, 0)
    # (correção numérica fina caso a planilha registre algo como 1e-15)
    if len(df) > 0:
        df.loc[0, ["ea", "q", "ev"]] = 0.0
        df.loc[0, "p"] = sigma3
        df.loc[0, "e_atual"] = E0_CONSTANTE

    # Calcular incrementos
    df["deltaEa"] = df["ea"].diff().fillna(0.0)
    df["deltaev"] = df["ev"].diff().fillna(0.0)
    df["dq"] = df["q"].diff().fillna(0.0)
    df["dp"] = df["p"].diff().fillna(0.0)

    # Adicionar identificadores
    df["po_kpa"] = sigma3
    df["teor_de_fibra"] = pf
    df["comp_de_fibra"] = L if L is not None else 0.0  # 0 quando sem fibra
    df["e_inicial_vazio_inicial"] = E0_CONSTANTE

    # Reorganizar para o formato padrão de 13 colunas
    df_out = df[[
        "po_kpa", "teor_de_fibra", "comp_de_fibra", "e_inicial_vazio_inicial",
        "p", "q", "e_atual",
        "ea", "ev",
        "deltaEa", "deltaev", "dq", "dp"
    ]].copy()
    df_out.columns = [
        "po_kpa", "teor_de_fibra", "comp_de_fibra", "e_inicial_vazio_inicial",
        "p_tensao_media", "q_tensao_desviadora", "e_indice_de_vazios",
        "ea_deformacao_axial", "ev_deformacao_volumetrica",
        "deltaEa", "deltaev", "dq", "dp"
    ]
    return df_out


# ============================================================================
# REMOÇÃO DE PADDING
# ============================================================================

def remover_padding(df):
    """
    Remove pontos onde a curva está estagnada (deltaEa ≈ 0 E dq ≈ 0).
    Mantém o primeiro ponto (que tem deltaEa=0 por construção).
    """
    if len(df) == 0:
        return df
    mask = (df["deltaEa"].abs() < 1e-10) & (df["dq"].abs() < 1e-10)
    mask.iloc[0] = False  # nunca remover o ponto inicial
    return df[~mask].reset_index(drop=True)


# ============================================================================
# PROCESSAMENTO DE UMA PLANILHA (12 ENSAIOS)
# ============================================================================

def processar_planilha(nome, path):
    """Lê a aba Exp_res e extrai os 12 ensaios experimentais."""
    print(f"\n{'='*70}")
    print(f"PLANILHA: {nome} ({path})")
    print(f"{'='*70}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Exp_res"]
    all_rows = list(ws.iter_rows(max_row=200, values_only=True))
    wb.close()

    if not all_rows or len(all_rows) < 4:
        print("✗ Planilha vazia ou sem cabeçalho.")
        return []

    # Padronizar comprimento das linhas
    max_cols = max(len(r) for r in all_rows)
    all_rows = [list(r) + [None] * (max_cols - len(r)) for r in all_rows]

    h1 = all_rows[0]  # σ3 e/ou Pf
    h2 = all_rows[1]  # Pf e/ou rótulo
    h3 = all_rows[2]  # L e/ou rótulo

    # Cada bloco de ensaio ocupa 4 colunas (3 dados + 1 vazia)
    # Iterar de 4 em 4 começando da coluna 0
    ensaios_extraidos = []
    j = 0
    while j < min(MAX_COL_EXP, max_cols):
        # Tentar parsear cabeçalho deste bloco
        # σ3 está geralmente em h1[j]
        # Pf pode estar em h1[j+1], h1[j], ou h2[j]
        # L pode estar em h1[j+2], h2[j], ou h3[j]
        sigma3 = parse_sigma3(h1[j])

        if sigma3 is None:
            # Coluna vazia — pular para próxima
            j += 1
            continue

        # Tentar Pf nas posições prováveis
        pf = parse_pf(h1[j+1] if j+1 < max_cols else None)
        if pf is None:
            pf = parse_pf(h2[j])

        # Tentar L nas posições prováveis
        L = parse_L(h1[j+2] if j+2 < max_cols else None)
        if L is None:
            L = parse_L(h3[j])
        if L is None:
            L = parse_L(h2[j])

        # Caso ensaio sem fibra: Pf=0 e L=None — manter L como None
        # (será setado para 0.0 ao salvar)
        if pf is None:
            pf = 0.0  # default para ensaios sem cabeçalho de Pf

        # Extrair ensaio
        df_ensaio = extrair_ensaio(all_rows, j, sigma3, pf, L, MAX_COL_EXP)

        if df_ensaio is not None and len(df_ensaio) >= 5:
            # Aplicar limpeza (remover padding)
            n_bruto = len(df_ensaio)
            df_limpo = remover_padding(df_ensaio)
            n_limpo = len(df_limpo)

            ensaios_extraidos.append({
                "fonte": nome,
                "col": j + 1,
                "sigma3": sigma3,
                "pf": pf,
                "L": L if L is not None else 0.0,
                "n_bruto": n_bruto,
                "n_limpo": n_limpo,
                "data": df_limpo,
            })

            L_str = f"{L:g}" if L is not None else "—"
            print(f"  Col {j+1:>3} | σ3={sigma3:>3} kPa | Pf={pf:.1f}% | L={L_str:>5} | "
                  f"{n_bruto} → {n_limpo} pontos ✓")

        # Avançar para próximo bloco (4 colunas adiante)
        j += 4

    print(f"\n>>> Total de ensaios extraídos de '{nome}': {len(ensaios_extraidos)}")
    return ensaios_extraidos


# ============================================================================
# PIPELINE PRINCIPAL
# ============================================================================

def main():
    print(f"Pipeline de extração — Tier 1")
    print(f"Premissa: e0 = {E0_CONSTANTE} (Dr = 60% constante, PINTO 2021)\n")

    # Processar as duas planilhas
    todos_ensaios = []
    for nome, path in PLANILHAS.items():
        ensaios = processar_planilha(nome, path)
        todos_ensaios.extend(ensaios)

    if not todos_ensaios:
        print("\n✗ Nenhum ensaio extraído. Abortando.")
        return

    # Verificar duplicatas (mesma combinação σ3, Pf, L em ambas as planilhas)
    print(f"\n{'='*70}")
    print("VERIFICAÇÃO DE DUPLICATAS")
    print(f"{'='*70}")
    combinacoes = {}
    for e in todos_ensaios:
        chave = (e["sigma3"], e["pf"], e["L"])
        combinacoes.setdefault(chave, []).append(e)

    duplicatas = {k: v for k, v in combinacoes.items() if len(v) > 1}
    if duplicatas:
        print(f"⚠ {len(duplicatas)} combinações aparecem em mais de uma planilha:")
        for k, lst in duplicatas.items():
            print(f"  σ3={k[0]}, Pf={k[1]}, L={k[2]}: presentes em "
                  f"{[e['fonte'] for e in lst]}")
    else:
        print("✓ Nenhuma duplicata — cada combinação é única.")

    # Concatenar tudo num DataFrame único
    df_final = pd.concat([e["data"] for e in todos_ensaios], ignore_index=True)
    print(f"\nTotal de pontos consolidados: {len(df_final)}")

    # Salvar CSV principal
    df_final.to_csv(CSV_SAIDA, index=False)
    print(f"✓ CSV salvo: {CSV_SAIDA}")

    # Salvar log
    log = pd.DataFrame([{
        "fonte": e["fonte"],
        "col_excel": e["col"],
        "sigma3_kPa": e["sigma3"],
        "Pf_pct": e["pf"],
        "L_mm": e["L"],
        "n_bruto": e["n_bruto"],
        "n_limpo": e["n_limpo"],
    } for e in todos_ensaios])
    log_path = CSV_SAIDA.replace(".csv", "_log.csv")
    log.to_csv(log_path, index=False)
    print(f"✓ Log salvo: {log_path}")

    # Estatísticas
    print(f"\n{'='*70}")
    print("ESTATÍSTICAS DO CONJUNTO EXPERIMENTAL")
    print(f"{'='*70}")
    print(f"\nDistribuição por (Pf, L):")
    print(df_final.groupby(["teor_de_fibra", "comp_de_fibra"]).size().to_string())
    print(f"\nDistribuição por σ3 (po_kpa):")
    print(df_final.groupby("po_kpa").size().to_string())

    print(f"\nFaixas das variáveis:")
    print(f"  εa máximo: {df_final['ea_deformacao_axial'].max():.4f} "
          f"({df_final['ea_deformacao_axial'].max()*100:.1f}%)")
    print(f"  q máximo:  {df_final['q_tensao_desviadora'].max():.1f} kPa")
    print(f"  p máximo:  {df_final['p_tensao_media'].max():.1f} kPa")
    print(f"  εv mín/máx: {df_final['ev_deformacao_volumetrica'].min():.4f} / "
          f"{df_final['ev_deformacao_volumetrica'].max():.4f}")

    # Comparação com o CSV processado anterior (se existir)
    try:
        df_anterior = pd.read_csv("dados_experimentais_processados.csv")
        n_ensaios_anterior = df_anterior.groupby(["po_kpa","teor_de_fibra","comp_de_fibra"]).ngroups
        n_ensaios_novo = df_final.groupby(["po_kpa","teor_de_fibra","comp_de_fibra"]).ngroups
        print(f"\n>>> Comparação com 'dados_experimentais_processados.csv' (artigo):")
        print(f"    Ensaios:  {n_ensaios_anterior} → {n_ensaios_novo} (+{n_ensaios_novo-n_ensaios_anterior})")
        print(f"    Pontos:   {len(df_anterior)} → {len(df_final)} (+{len(df_final)-len(df_anterior)})")
    except FileNotFoundError:
        pass


if __name__ == "__main__":
    main()
